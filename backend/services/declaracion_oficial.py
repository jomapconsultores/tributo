"""Llena el formulario oficial del SRI (IVA / ICE) a partir de la declaración
calculada. Criterio: el valor se escribe en la celda contigua (a la derecha) de
cada código SRI. Se omiten los códigos de la sección RESULTADO (399, 499…) que
el propio formulario calcula con sus fórmulas. Es un borrador a verificar."""
import io
import os
import openpyxl

TEMPLATES = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "resources", "templates")

# Mapeo de los casilleros INTERNOS del sistema a los OFICIALES del formulario
# 104 del SRI (algunos no coinciden: 412 interno = ventas 5%, pero 412 oficial
# = activos fijos). Solo se mapean las entradas (ventas/adquisiciones); el
# formulario calcula el resto. Un valor None = no se traslada.
MAP_IVA = {
    # Ventas
    "411": "411",   # 15% neto
    "421": "421",   # 15% IVA
    "412": "420",   # 5% neto   (oficial 420)
    "422": "430",   # 5% IVA    (oficial 430)
    "413": "413",   # 0% neto
    "414": None,    # exentas (no hay casillero de ventas estándar)
    "415": "441",   # no objeto neto
    # Adquisiciones
    "510": "510", "520": "520",   # 15% neto / IVA
    "550": "550", "560": "560",   # 5% neto / IVA
    "517": "517",                 # 0% neto
    "518": "541",                 # no objeto neto
    "519": "542",                 # exentas neto
}


def llenar_oficial(tipo, decl):
    es_ice = str(tipo).upper() == "ICE"
    fname = "ice_form.xlsx" if es_ice else "iva_form.xlsx"
    path = os.path.join(TEMPLATES, fname)
    if not os.path.exists(path):
        raise FileNotFoundError("No está la plantilla oficial: " + fname)

    # código -> valor (solo entradas, no resultados/fórmulas)
    valores = {}
    # La plantilla del SRI es de DISEÑO (sin fórmulas): se llenan TODOS los
    # casilleros que calculamos, incluido el resultado (crédito mes anterior
    # 605/606, factor 563, crédito 564, etc.). Las celdas de VENTAS/
    # ADQUISICIONES se traducen al casillero oficial; las de RESULTADO ya
    # usan el código oficial.
    # Códigos internos SIN casillero oficial estándar del SRI (p. ej. "R-50"
    # rebaja ICE, "EXE" exención ICE — Art. 82/77.1 LRTI): no hay un casillero
    # de la sección RESULTADO del formulario 104 al que trasladarlos de forma
    # confiable, así que NO se descartan en silencio. Si tienen valor, se
    # reportan en "omitidos" para que el contador los ubique manualmente.
    no_estandar = {}
    for f in decl.get("filas", []):
        cod = str(f.get("codigo", "")).strip()
        if not cod.isdigit():
            valor = f.get("valor", 0)
            if valor:
                no_estandar[cod] = valor
            continue
        if not es_ice and f.get("seccion") != "RESULTADO":
            cod = MAP_IVA.get(cod, cod)   # traducir al casillero oficial
            if cod is None:
                continue
        valores[cod] = f.get("valor", 0)

    wb = openpyxl.load_workbook(path)
    llenados, omitidos = [], []
    for cod, valor in no_estandar.items():
        omitidos.append(f"{cod}={valor} (sin casillero oficial del SRI — ubicar manualmente)")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                s = str(cell.value).strip()
                if s not in valores:
                    continue
                target = ws.cell(row=cell.row, column=cell.column + 1)
                tv = target.value
                if isinstance(tv, str) and tv.startswith("="):
                    omitidos.append(s)
                    continue
                try:
                    target.value = valores[s]   # MergedCell de solo lectura lanza excepción
                    llenados.append(s)
                except Exception:
                    omitidos.append(s)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue(), sorted(set(llenados)), sorted(set(omitidos))
