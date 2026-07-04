import csv
import io
import re
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from typing import Optional, List
from pydantic import BaseModel
from auth import get_current_user
from database import get_supabase_client
from tenancy import can_access_identificacion
from services.min_produccion import verificar_ruc
from services.doc_ia import leer_documento_ia

router = APIRouter(prefix="/api/rebajas", tags=["rebajas"])


def _owner_uid(supabase, identificacion: str, user_id: str) -> str:
    """Resuelve el user_id DUEÑO real del contribuyente (el mismo para todos sus
    períodos). Los componentes/proveedores de rebajas se guardan SIEMPRE bajo el
    dueño real (no bajo quien los carga) para que _rebajas_por_producto (usada al
    calcular la declaración ICE, que consulta por owner_uid) los encuentre sin
    importar si quien los cargó fue un socio con acceso compartido."""
    if not can_access_identificacion(user_id, identificacion):
        raise HTTPException(status_code=404, detail="Contribuyente no encontrado")
    r = supabase.table("clients").select("user_id").eq("identificacion", identificacion).limit(1).execute().data
    return r[0]["user_id"] if r else user_id


def _owner_uid_de_fila(supabase, tabla: str, rid: str, user_id: str) -> str:
    """Para endpoints con :id (sin identificacion en la URL): ubica la fila,
    verifica que el usuario puede acceder a su identificacion, y devuelve el
    owner_uid real de esa fila (para reusar el mismo filtro de guardado)."""
    row = supabase.table(tabla).select("identificacion,user_id").eq("id", rid).execute().data
    if not row:
        raise HTTPException(status_code=404, detail="No encontrado")
    ident = row[0].get("identificacion")
    if not can_access_identificacion(user_id, ident):
        raise HTTPException(status_code=404, detail="No encontrado")
    return row[0].get("user_id") or user_id


@router.get("/verificar-ruc")
async def verificar(ruc: str = Query(...), _: str = Depends(get_current_user)):
    """Verifica en el Ministerio de Producción si el RUC está categorizado."""
    return verificar_ruc(ruc)

COLUMNS = "id,identificacion,producto,ingrediente,ruc_proveedor,proveedor_nombre,cantidad,unidad,densidad,origen,calificado"


class RebajaIn(BaseModel):
    identificacion: str
    producto: str
    ingrediente: str
    ruc_proveedor: Optional[str] = ""
    proveedor_nombre: Optional[str] = ""
    cantidad: float = 0
    unidad: Optional[str] = "ml"
    densidad: Optional[float] = 1
    origen: Optional[str] = "NACIONAL"
    calificado: Optional[bool] = False


@router.get("/")
async def list_rebajas(
    identificacion: str = Query(...),
    producto: Optional[str] = Query(None),
    user_id: str = Depends(get_current_user),
):
    try:
        supabase = get_supabase_client()
        owner_uid = _owner_uid(supabase, identificacion, user_id)
        q = supabase.table("rebajas_ingredientes").select(COLUMNS).eq("identificacion", identificacion).eq("user_id", owner_uid)
        if producto:
            q = q.eq("producto", producto)
        return {"data": q.order("ingrediente").execute().data or []}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/")
async def create_rebaja(entry: RebajaIn, user_id: str = Depends(get_current_user)):
    try:
        supabase = get_supabase_client()
        data = entry.dict()
        data["user_id"] = _owner_uid(supabase, entry.identificacion, user_id)
        data["ingrediente"] = (data.get("ingrediente") or "").strip().upper()
        data["origen"] = (data.get("origen") or "NACIONAL").upper()
        if not data["ingrediente"]:
            raise HTTPException(status_code=400, detail="El ingrediente es obligatorio")
        res = supabase.table("rebajas_ingredientes").insert(data).execute()
        return res.data[0] if res.data else None
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class RebajaUpdate(BaseModel):
    ingrediente: Optional[str] = None
    ruc_proveedor: Optional[str] = None
    proveedor_nombre: Optional[str] = None
    cantidad: Optional[float] = None
    unidad: Optional[str] = None
    densidad: Optional[float] = None
    calificado: Optional[bool] = None


@router.put("/{rid}")
async def update_rebaja(rid: str, entry: RebajaUpdate, user_id: str = Depends(get_current_user)):
    """Edita un componente (p. ej. asignar/cambiar el RUC del proveedor por fila)."""
    try:
        supabase = get_supabase_client()
        owner_uid = _owner_uid_de_fila(supabase, "rebajas_ingredientes", rid, user_id)
        data = {k: v for k, v in entry.dict().items() if v is not None}
        if "ingrediente" in data:
            data["ingrediente"] = data["ingrediente"].strip().upper()
        if "proveedor_nombre" in data:
            data["proveedor_nombre"] = (data["proveedor_nombre"] or "").strip().upper()
        res = supabase.table("rebajas_ingredientes").update(data).eq("id", rid).eq("user_id", owner_uid).execute()
        return res.data[0] if res.data else None
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/{rid}")
async def delete_rebaja(rid: str, user_id: str = Depends(get_current_user)):
    try:
        supabase = get_supabase_client()
        owner_uid = _owner_uid_de_fila(supabase, "rebajas_ingredientes", rid, user_id)
        supabase.table("rebajas_ingredientes").delete().eq("id", rid).eq("user_id", owner_uid).execute()
        return {"message": "Eliminado"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ── Carga masiva de componentes (pegar de Excel / subir archivo) ──

class RebajaBulkItem(BaseModel):
    ingrediente: str
    ruc_proveedor: Optional[str] = ""
    proveedor_nombre: Optional[str] = ""
    cantidad: float = 0
    unidad: Optional[str] = "ml"
    densidad: Optional[float] = 1
    origen: Optional[str] = "NACIONAL"
    calificado: Optional[bool] = False


class RebajaBulkIn(BaseModel):
    identificacion: str
    producto: str
    items: List[RebajaBulkItem]


def _f(v, d=0.0):
    try:
        return float(str(v).replace(",", ".").strip())
    except (TypeError, ValueError):
        return d


def _split_cantidad(raw, unidad_col=None):
    """Separa la cantidad de su tipo de medida: '700 ml' -> (700.0, 'ml').
    Si la unidad viene en columna aparte, se respeta."""
    u = (str(unidad_col or "")).strip()
    s = str(raw or "").strip().replace(",", ".")
    m = re.match(r"^\s*([0-9]*\.?[0-9]+)\s*([a-zA-Zµ]+)?\s*$", s)
    if m:
        num = _f(m.group(1))
        if not u and m.group(2):
            u = m.group(2)
    else:
        num = _f(s)
    return num, (u or "ml")


def _insertar_componentes(supabase, user_id, identificacion, producto, items):
    """Inserta una lista de componentes; autocompleta calificado/nombre desde el
    catálogo de proveedores si el RUC ya está guardado. Devuelve cuántos insertó."""
    producto = (producto or "").strip().upper()
    # Catálogo de proveedores del contribuyente (RUC → nombre/calificado)
    cat = {}
    try:
        pv = supabase.table("rebajas_proveedores").select("ruc,nombre,calificado")\
            .eq("user_id", user_id).eq("identificacion", identificacion).execute().data or []
        cat = {p["ruc"]: p for p in pv}
    except Exception:
        cat = {}
    filas = []
    for it in items:
        ing = (it.get("ingrediente") if isinstance(it, dict) else it.ingrediente) or ""
        ing = ing.strip().upper()
        if not ing:
            continue
        d = it if isinstance(it, dict) else it.dict()
        ruc = (d.get("ruc_proveedor") or "").strip()
        prov = cat.get(ruc)
        qty, uni = _split_cantidad(d.get("cantidad"), d.get("unidad"))
        filas.append({
            "user_id": user_id, "identificacion": identificacion, "producto": producto,
            "ingrediente": ing, "ruc_proveedor": ruc,
            "proveedor_nombre": (d.get("proveedor_nombre") or (prov or {}).get("nombre") or "").strip().upper(),
            "cantidad": qty, "unidad": uni,
            "densidad": _f(d.get("densidad"), 1) or 1,
            "origen": (d.get("origen") or "NACIONAL").upper(),
            "calificado": bool(d.get("calificado")) or bool((prov or {}).get("calificado")),
        })
    if not filas:
        return 0
    supabase.table("rebajas_ingredientes").insert(filas).execute()
    return len(filas)


@router.post("/bulk")
async def bulk_create(entry: RebajaBulkIn, user_id: str = Depends(get_current_user)):
    """Carga masiva de componentes pegados desde Excel (filas ya parseadas)."""
    try:
        supabase = get_supabase_client()
        owner_uid = _owner_uid(supabase, entry.identificacion, user_id)
        n = _insertar_componentes(supabase, owner_uid, entry.identificacion, entry.producto, entry.items)
        return {"insertados": n}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


_COL_ALIASES = {
    "ruc_proveedor": ("ruc", "ruc proveedor", "ruc_proveedor", "ruc del proveedor"),
    "proveedor_nombre": ("proveedor", "empresa", "nombre", "razon social", "razón social", "proveedor_nombre"),
    "ingrediente": ("ingrediente", "componente", "producto", "materia prima", "insumo"),
    "cantidad": ("cantidad", "cant", "qty", "volumen", "peso"),
    "unidad": ("unidad", "und", "um", "u.m."),
    "densidad": ("densidad", "dens", "densidad g/ml"),
}


def _norm(s):
    return str(s or "").strip().lower()


def _mapear_encabezados(header):
    idx = {}
    for j, h in enumerate(header):
        hn = _norm(h)
        for campo, alias in _COL_ALIASES.items():
            if campo not in idx and hn in alias:
                idx[campo] = j
    return idx


def _filas_a_items(rows):
    """rows: lista de listas (primera fila = encabezados). Devuelve items dict."""
    if not rows:
        return []
    idx = _mapear_encabezados(rows[0])
    if "ingrediente" not in idx:
        raise HTTPException(status_code=400, detail="No se encontró la columna de ingrediente/componente. Encabezados esperados: RUC, Proveedor, Ingrediente, Cantidad, Unidad, Densidad.")
    items = []
    for r in rows[1:]:
        def g(c):
            j = idx.get(c)
            return r[j] if (j is not None and j < len(r)) else ""
        if not _norm(g("ingrediente")):
            continue
        items.append({
            "ruc_proveedor": str(g("ruc_proveedor") or "").strip(),
            "proveedor_nombre": str(g("proveedor_nombre") or "").strip(),
            "ingrediente": str(g("ingrediente") or "").strip(),
            # Unidad vacía si no hay columna: _split_cantidad la extrae de la cantidad ("50 g")
            "cantidad": g("cantidad"), "unidad": str(g("unidad") or "").strip(),
            "densidad": g("densidad"),
        })
    return items


@router.post("/parse-file")
async def parse_file(
    file: UploadFile = File(...),
    identificacion: str = Form(...),
    producto: str = Form(...),
    user_id: str = Depends(get_current_user),
):
    """Sube un .xlsx/.csv de componentes, lo parsea e inserta. Devuelve cuántos."""
    try:
        content = await file.read()
        name = (file.filename or "").lower()
        rows = []
        if name.endswith(".csv") or name.endswith(".txt"):
            text = content.decode("utf-8-sig", errors="replace")
            sep = "\t" if "\t" in text.splitlines()[0] else ","
            rows = [r for r in csv.reader(io.StringIO(text), delimiter=sep)]
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            sh = wb.active
            for r in sh.iter_rows(values_only=True):
                rows.append(list(r))
        items = _filas_a_items(rows)
        supabase = get_supabase_client()
        owner_uid = _owner_uid(supabase, identificacion, user_id)
        n = _insertar_componentes(supabase, owner_uid, identificacion, producto, items)
        return {"insertados": n}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {e}")


# ── Catálogo reutilizable de proveedores (RUC → nombre + calificado) ──

PROV_COLS = "id,identificacion,ruc,nombre,calificado,categoria,actividad,vigencia,vigencia_inicio,vigente_hasta,documentos,verificado_at"
PROV_BUCKET = "proveedores"


def _texto_de_archivo(content, filename):
    """Extrae texto plano de un PDF/Excel/CSV para buscar datos (RUC). Las fotos
    no se leen (no hay OCR)."""
    name = (filename or "").lower()
    try:
        if name.endswith(".pdf"):
            import PyPDF2
            rd = PyPDF2.PdfReader(io.BytesIO(content))
            return "\n".join((p.extract_text() or "") for p in rd.pages)
        if name.endswith(".csv") or name.endswith(".txt"):
            return content.decode("utf-8-sig", errors="replace")
        if name.endswith(".xlsx") or name.endswith(".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            out = []
            for sh in wb.worksheets:
                for row in sh.iter_rows(values_only=True):
                    out.append(" ".join("" if c is None else str(c) for c in row))
            return "\n".join(out)
    except Exception as e:
        print(f"_texto_de_archivo: {e}")
    return ""


def _extraer_ruc(content, filename):
    """Busca un RUC (13 dígitos terminado en 001) en el texto del documento."""
    txt = _texto_de_archivo(content, filename)
    if not txt:
        return None
    m = re.findall(r"\b(\d{13})\b", txt)
    if not m:
        return None
    # Preferir los que terminan en 001 (RUC de sociedad/persona con establecimiento)
    for r in m:
        if r.endswith("001"):
            return r
    return m[0]


class ProveedorIn(BaseModel):
    identificacion: str
    ruc: str
    nombre: Optional[str] = ""
    calificado: Optional[bool] = False
    categoria: Optional[str] = ""
    actividad: Optional[str] = ""
    vigencia: Optional[str] = ""
    vigencia_inicio: Optional[str] = None
    vigente_hasta: Optional[str] = None


@router.get("/proveedores")
async def list_proveedores(identificacion: str = Query(...), user_id: str = Depends(get_current_user)):
    try:
        supabase = get_supabase_client()
        owner_uid = _owner_uid(supabase, identificacion, user_id)
        r = supabase.table("rebajas_proveedores").select(PROV_COLS)\
            .eq("identificacion", identificacion).eq("user_id", owner_uid).order("nombre").execute()
        return {"data": r.data or []}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/proveedores/enriquecer-actividades")
async def enriquecer_actividades_prov(identificacion: str = Query(...), user_id: str = Depends(get_current_user)):
    """Trae automáticamente la actividad económica del SRI para los proveedores
    calificados que aún no la tienen. Por lotes; el frontend llama en bucle."""
    try:
        from services.min_produccion import consultar_sri
        supabase = get_supabase_client()
        owner_uid = _owner_uid(supabase, identificacion, user_id)
        rows = supabase.table("rebajas_proveedores").select("id,ruc,actividad")\
            .eq("identificacion", identificacion).eq("user_id", owner_uid).execute().data or []
        faltan = [r for r in rows if (r.get("ruc") or "").strip() and not (r.get("actividad") or "").strip()]
        lote = faltan[:8]
        actualizados = 0
        for r in lote:
            ruc = (r.get("ruc") or "").strip()
            try:
                sri = consultar_sri(ruc, timeout=6) or {}
            except Exception:
                sri = {}
            ae = (sri.get("actividad_economica") or "").strip() or "—"
            try:
                supabase.table("rebajas_proveedores").update({"actividad": ae}).eq("id", r["id"]).execute()
                if ae != "—":
                    actualizados += 1
            except Exception:
                pass
        return {"actualizados": actualizados, "procesados": len(lote), "restantes": max(0, len(faltan) - len(lote))}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _upsert_proveedor(supabase, user_id, ident, ruc, nombre, calificado, categoria="", vigencia="", vigente_hasta=None, vigencia_inicio=None, actividad=None):
    data = {
        "user_id": user_id, "identificacion": ident, "ruc": (ruc or "").strip(),
        "nombre": (nombre or "").strip().upper(), "calificado": bool(calificado),
        "categoria": categoria or "", "vigencia": vigencia or "",
        "verificado_at": datetime.now(timezone.utc).isoformat(),
    }
    if vigente_hasta:
        data["vigente_hasta"] = vigente_hasta
    if vigencia_inicio:
        data["vigencia_inicio"] = vigencia_inicio
    if actividad:
        data["actividad"] = actividad
    return supabase.table("rebajas_proveedores").upsert(
        data, on_conflict="user_id,identificacion,ruc").execute()


@router.put("/proveedores")
async def upsert_proveedor(entry: ProveedorIn, user_id: str = Depends(get_current_user)):
    """Guarda/actualiza un proveedor en el catálogo (tras verificar el RUC)."""
    try:
        if not (entry.ruc or "").strip():
            raise HTTPException(status_code=400, detail="El RUC es obligatorio")
        supabase = get_supabase_client()
        owner_uid = _owner_uid(supabase, entry.identificacion, user_id)
        res = _upsert_proveedor(supabase, owner_uid, entry.identificacion, entry.ruc,
                                entry.nombre, entry.calificado, entry.categoria, entry.vigencia, entry.vigente_hasta, entry.vigencia_inicio, entry.actividad)
        return res.data[0] if res.data else None
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/proveedores/{pid}")
async def delete_proveedor(pid: str, user_id: str = Depends(get_current_user)):
    try:
        supabase = get_supabase_client()
        owner_uid = _owner_uid_de_fila(supabase, "rebajas_proveedores", pid, user_id)
        supabase.table("rebajas_proveedores").delete().eq("id", pid).eq("user_id", owner_uid).execute()
        return {"message": "Eliminado"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


def _prov_bucket(supabase):
    try:
        names = {getattr(b, "name", None) or (b.get("name") if isinstance(b, dict) else None) for b in supabase.storage.list_buckets()}
        if PROV_BUCKET not in names:
            supabase.storage.create_bucket(PROV_BUCKET, options={"public": False})
    except Exception as e:
        print(f"_prov_bucket: {e}")


@router.post("/proveedores/documento")
async def subir_documento(
    file: UploadFile = File(...),
    identificacion: str = Form(...),
    ruc: Optional[str] = Form(None),
    nombre: Optional[str] = Form(""),
    calificado: Optional[bool] = Form(False),
    vigente_hasta: Optional[str] = Form(None),
    user_id: str = Depends(get_current_user),
):
    """Sube un documento (PDF/Excel) y EXTRAE los datos del proveedor: si no se
    envía el RUC, lo lee del documento, consulta el Ministerio y registra nombre,
    calificación y vigencia (inicio–fin) automáticamente."""
    try:
        supabase = get_supabase_client()
        owner_uid = _owner_uid(supabase, identificacion, user_id)
        _prov_bucket(supabase)
        content = await file.read()
        auto = not (ruc or "").strip()
        # 1) Lectura con IA (lee fotos, escaneos, PDF y Excel)
        ia = {}
        if auto:
            low = (file.filename or "").lower()
            texto = None if (low.endswith(".pdf") or low.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif"))) else _texto_de_archivo(content, file.filename)
            ia = leer_documento_ia(content, file.filename, file.content_type, texto) or {}
            ruc = ia.get("ruc") or _extraer_ruc(content, file.filename)
            if not ruc:
                raise HTTPException(status_code=400, detail="No se pudo leer el RUC del documento. Verifica que el documento muestre el RUC, o escríbelo manualmente. (Para lectura de fotos/escaneos se requiere MISTRAL_API_KEY configurada.)")
        ruc = ruc.strip()
        # 2) Respaldo Ministerio si la IA no obtuvo categoría/vigencia
        verif = {}
        falta = not (ia.get("categoria") and (ia.get("vigencia_fin") or ia.get("vigencia_inicio")))
        if (auto and falta) or (not auto and not nombre):
            try:
                verif = verificar_ruc(ruc)
            except Exception:
                verif = {}
        safe = re.sub(r"[^A-Za-z0-9._-]", "_", file.filename or "documento")
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
        path = f"{identificacion}/{ruc}/{stamp}_{safe}"
        supabase.storage.from_(PROV_BUCKET).upload(
            path, content, {"content-type": file.content_type or "application/octet-stream", "upsert": "true"})
        cur = supabase.table("rebajas_proveedores").select("id,documentos,nombre,calificado,vigente_hasta,vigencia_inicio,actividad")\
            .eq("user_id", owner_uid).eq("identificacion", identificacion).eq("ruc", ruc).execute().data
        prev = cur[0] if cur else {}
        docs = (prev.get("documentos") or []) if prev else []
        docs.append({"nombre": file.filename, "path": path, "subido": datetime.now(timezone.utc).isoformat()})
        data = {
            "user_id": owner_uid, "identificacion": identificacion, "ruc": ruc,
            "nombre": (nombre or ia.get("nombre") or verif.get("razon_social") or prev.get("nombre") or "").strip().upper(),
            "calificado": bool(calificado) or bool(ia.get("calificado")) or bool(verif.get("cumple")) or bool(prev.get("calificado")),
            "categoria": ia.get("categoria") or verif.get("categoria") or "",
            "actividad": verif.get("actividad_economica") or prev.get("actividad") or "",
            "vigencia": verif.get("vigencia") or "",
            "documentos": docs,
            "verificado_at": datetime.now(timezone.utc).isoformat(),
        }
        vi = ia.get("vigencia_inicio") or verif.get("vigencia_inicio")
        vf = vigente_hasta or ia.get("vigencia_fin") or verif.get("vigencia_fin")
        if vi:
            data["vigencia_inicio"] = vi
        if vf:
            data["vigente_hasta"] = vf
        res = supabase.table("rebajas_proveedores").upsert(data, on_conflict="user_id,identificacion,ruc").execute()
        return res.data[0] if res.data else None
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo subir el documento: {e}")


@router.get("/proveedores/documento-url")
async def documento_url(path: str = Query(...), user_id: str = Depends(get_current_user)):
    """Devuelve una URL firmada temporal para ver/descargar un documento."""
    try:
        # El path se guarda como "{identificacion}/{ruc}/{archivo}" (ver
        # subir_documento) — valida acceso al RUC antes de firmar la URL, para
        # que un usuario no pueda descargar documentos de otro contribuyente
        # solo por adivinar/conocer el path.
        identificacion = path.split("/", 1)[0] if path else ""
        if not identificacion or not can_access_identificacion(user_id, identificacion):
            raise HTTPException(status_code=404, detail="No encontrado")
        supabase = get_supabase_client()
        r = supabase.storage.from_(PROV_BUCKET).create_signed_url(path, 3600)
        url = r.get("signedURL") or r.get("signedUrl") or r.get("signed_url") if isinstance(r, dict) else None
        return {"url": url}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/proveedores/verificar-todos")
async def verificar_todos(
    identificacion: str = Query(...),
    producto: Optional[str] = Query(None),
    user_id: str = Depends(get_current_user),
):
    """Verifica en el Ministerio todos los RUC distintos (de un producto o del RUC
    completo), actualiza el catálogo de proveedores y los componentes guardados."""
    try:
        supabase = get_supabase_client()
        owner_uid = _owner_uid(supabase, identificacion, user_id)
        q = supabase.table("rebajas_ingredientes").select("ruc_proveedor")\
            .eq("identificacion", identificacion).eq("user_id", owner_uid)
        if producto:
            q = q.eq("producto", producto.strip().upper())
        rucs = sorted({(r.get("ruc_proveedor") or "").strip() for r in (q.execute().data or [])} - {""})
        resultados = []
        for ruc in rucs:
            try:
                d = verificar_ruc(ruc)
            except Exception as e:
                resultados.append({"ruc": ruc, "error": str(e)})
                continue
            cumple = bool(d.get("cumple"))
            nombre = d.get("razon_social") or ""
            _upsert_proveedor(supabase, owner_uid, identificacion, ruc, nombre, cumple,
                              d.get("categoria", ""), d.get("vigencia", ""),
                              d.get("vigencia_fin"), d.get("vigencia_inicio"), d.get("actividad_economica"))
            # Propaga a los componentes guardados con ese RUC
            upd = supabase.table("rebajas_ingredientes").update(
                {"calificado": cumple, "proveedor_nombre": (nombre or "").upper()})\
                .eq("identificacion", identificacion).eq("user_id", owner_uid).eq("ruc_proveedor", ruc)
            if producto:
                upd = upd.eq("producto", producto.strip().upper())
            upd.execute()
            resultados.append({"ruc": ruc, "cumple": cumple, "nombre": nombre,
                               "categoria": d.get("categoria", ""), "mensaje": d.get("mensaje", "")})
        return {"verificados": len(resultados), "data": resultados}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Condiciones normativas por producto (Art. 82/77 LRTI, Art. 199.4/199.5 RLRTI) ──

PROD_COLS = "id,identificacion,producto,es_cerveza,nueva_marca,cupo_anual_sri"


class CondicionesProducto(BaseModel):
    identificacion: str
    producto: str
    es_cerveza: bool = False        # cerveza: rebaja/exención solo para nuevas marcas
    nueva_marca: bool = False       # sin marca primigenia + nueva notificación sanitaria
    cupo_anual_sri: bool = False    # cupo anual del SRI (requisito de la exención)


@router.get("/producto")
async def get_condiciones(
    identificacion: str = Query(...),
    producto: Optional[str] = Query(None),
    user_id: str = Depends(get_current_user),
):
    """Condiciones normativas guardadas (de un producto, o todas las del RUC)."""
    try:
        supabase = get_supabase_client()
        owner_uid = _owner_uid(supabase, identificacion, user_id)
        q = supabase.table("rebajas_productos").select(PROD_COLS).eq(
            "identificacion", identificacion).eq("user_id", owner_uid)
        if producto:
            q = q.eq("producto", producto.strip().upper())
        return {"data": q.execute().data or []}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/producto")
async def set_condiciones(entry: CondicionesProducto, user_id: str = Depends(get_current_user)):
    """Crea o actualiza las condiciones normativas del producto (upsert)."""
    try:
        supabase = get_supabase_client()
        data = entry.dict()
        data["producto"] = (data.get("producto") or "").strip().upper()
        if not data["producto"]:
            raise HTTPException(status_code=400, detail="El producto es obligatorio")
        data["user_id"] = _owner_uid(supabase, entry.identificacion, user_id)
        res = supabase.table("rebajas_productos").upsert(
            data, on_conflict="user_id,identificacion,producto").execute()
        return res.data[0] if res.data else None
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
